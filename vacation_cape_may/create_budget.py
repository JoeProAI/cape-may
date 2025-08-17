import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def create_budget_tracker():
    # Read the activities CSV
    try:
        activities_df = pd.read_csv('data/activities.csv')
    except FileNotFoundError:
        print("Error: data/activities.csv not found.")
        return

    # Create a new Excel workbook
    wb = Workbook()
    
    # Activities Sheet
    ws_activities = wb.active
    ws_activities.title = "activities"
    
    # Filter and prepare activities data
    budget_activities = activities_df[activities_df['cost_estimate'].str.contains("\$", na=False)].copy()
    budget_activities['est_cost'] = budget_activities['cost_estimate'].str.extract(r'\\\$(\d+)').astype(float)
    budget_activities_display = budget_activities[['name', 'est_cost']]
    budget_activities_display['count'] = 1 # Default count
    budget_activities_display['notes'] = ''

    for r in dataframe_to_rows(budget_activities_display, index=False, header=True):
        ws_activities.append(r)

    # Daily Sheet
    ws_daily = wb.create_sheet(title="daily")
    ws_daily.append(['date', 'subtotal_activities', 'meals', 'misc'])
    for i in range(17, 24):
        ws_daily.append([f'2025-08-{i}', 0, 0, 0])

    # Summary Sheet
    ws_summary = wb.create_sheet(title="summary")
    ws_summary.append(['category', 'total'])
    ws_summary.append(['Total Activities', '=SUM(activities!B2:B100)'])
    ws_summary.append(['Total Meals', '=SUM(daily!C2:C8)'])
    ws_summary.append(['Total Misc', '=SUM(daily!D2:D8)'])
    ws_summary.append(['Grand Total', '=SUM(B2:B4)'])
    ws_summary.append(['Per Adult (6)', '=B5/6'])
    ws_summary.append(['Per Kid (3)', '=B5/3'])

    # Save the workbook
    wb.save('budget/budget.xlsx')
    print("Budget tracker created successfully.")

if __name__ == "__main__":
    create_budget_tracker()